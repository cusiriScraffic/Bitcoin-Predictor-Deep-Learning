from __future__ import unicode_literals, print_function, division
from io import open
import glob
import unicodedata
import string
import torch
import torch.nn as nn
from torch.autograd import Variable
import openpyxl
import random
import torch.nn.functional as F
from pprint import pprint
import torch.optim as optim
import numpy as np
import pandas as pd

workbook = openpyxl.load_workbook('/Users/calvinusiri/Desktop/CS/machinelearning/cryptoPrice1.xlsx')
type(workbook)
dictionaryOfCryptoObjects = dict()
sheetNames = ["BTC","ETH","XRP","LTC"]
crypto = "BTC"
numberOfOutChanels = 32

#Single input is 6x32 because 16 dilations
class Crypto(object):
	open = []
	high = []
	low = []
	close = []
	volume = []
	mktCap = []
	date = []
	def __init__(self,open,high,low,close,volume,mktCap,date):
		super(Crypto, self).__init__()
		self.open = open
		self.high = high
		self.low = low
		self.close = close
		self.volume = volume
		self.mktCap = mktCap
		self.date = date

class Net(nn.Module):
	def __init__(self):
		super(Net, self).__init__()
		self.conv = nn.Conv2d(in_channels=5,out_channels=numberOfOutChanels,kernel_size=(1,2),stride=1,dilation=1).double()
		self.conv1 = nn.Conv2d(in_channels=numberOfOutChanels,out_channels=numberOfOutChanels,kernel_size=(1,2),stride=1,dilation=2).double()
		self.conv2 = nn.Conv2d(in_channels=numberOfOutChanels,out_channels=numberOfOutChanels,kernel_size=(1,2),stride=1,dilation=4).double()
		self.conv3 = nn.Conv2d(in_channels=numberOfOutChanels,out_channels=numberOfOutChanels,kernel_size=(1,2),stride=1,dilation=8).double()
		self.conv4 = nn.Conv2d(in_channels=numberOfOutChanels,out_channels=numberOfOutChanels,kernel_size=(1,2),stride=1,dilation=16).double()
		self.fc1 = nn.Linear(128,1).double()

	def forward(self,x):
		x = F.relu(self.conv(x))
		x = F.relu(self.conv1(x))
		x = F.relu(self.conv2(x))
		x = F.relu(self.conv3(x))
		x = F.relu(self.conv4(x))
		x = x.view(64,128)
		x = self.fc1(x)
		x = nn.functional.tanh(x)
		return x

def getPercentageChange(arrays):
	container = []
	for a, b in zip(arrays[::1], arrays[1::1]):
		container.append(((b - a) / a)) # Removed 100* because, we want numbers close to 0 for sigmoid function
	return container

def reverseList(arr):
	arr.reverse()
	return arr

for i in sheetNames:
	sheet = workbook.get_sheet_by_name(i)
	open = []
	high = []
	low = []
	close = []
	volume = []
	mktCap = []
	date = []
	for j in range(1,sheet.max_column+1):
		value = sheet.cell(row=1,column=j).value
		if value == "Date":
			for k in range(2,907):#sheet.max_row+1): 
				date.append(sheet.cell(row=k,column=j).value)
		if value == "Open":
			for k in range(2,907):#sheet.max_row+1):
				open.append(sheet.cell(row=k,column=j).value)
		if value == "High":
			for k in range(2,907):#sheet.max_row+1):
				high.append(sheet.cell(row=k,column=j).value)
		if value == "Low":
			for k in range(2,907):#sheet.max_row+1):
				low.append(sheet.cell(row=k,column=j).value)
		if value == "Close":
			for k in range(2,907):#sheet.max_row+1):
				close.append(sheet.cell(row=k,column=j).value)
		if value == "Volume":
			for k in range(2,907):#sheet.max_row+1):
				value = sheet.cell(row=k,column=j).value
				if value == "-":
					volume.append(0.5)
				else:
					volume.append(value)
		if value == "Market Cap":
			for k in range(2,907):#sheet.max_row+1):
				mktCap.append(sheet.cell(row=k,column=j).value)
	dictionaryOfCryptoObjects[i] = Crypto(reverseList(open),reverseList(high),reverseList(low),reverseList(close),reverseList(volume),reverseList(mktCap),reverseList(date))

lengthOfCloseData = len(dictionaryOfCryptoObjects[crypto].close)
trainingData = int(lengthOfCloseData*0.80)
batchSize = 64

#Training network
cnn = Net()
criterion = nn.L1Loss()
optimizer = optim.SGD(cnn.parameters(), lr=0.001, momentum=0.9)
matrix = np.zeros((5,4,lengthOfCloseData-1))# Create matrix of data
for i in range(len(sheetNames)):
	matrix[0,i,:] = getPercentageChange(dictionaryOfCryptoObjects[sheetNames[i]].open)
	matrix[1,i,:] = getPercentageChange(dictionaryOfCryptoObjects[sheetNames[i]].close)
	matrix[2,i,:] = getPercentageChange(dictionaryOfCryptoObjects[sheetNames[i]].high)
	matrix[3,i,:] = getPercentageChange(dictionaryOfCryptoObjects[sheetNames[i]].low)
	matrix[4,i,:] = getPercentageChange(dictionaryOfCryptoObjects[sheetNames[i]].mktCap)

for sample in range(1000):
	input = np.zeros((batchSize,5,4,32))
	target = np.zeros((batchSize,1))
	for batch in range(batchSize):
		start = np.random.randint(trainingData)
		singleInputBatching = matrix[:,:,start:start+32]
		input[batch,:,:,:] = singleInputBatching
		target[batch,0] = matrix[1,3,start+32] #close,Which coin to predict,the next day
	input = torch.from_numpy(input).type(torch.DoubleTensor)
	cnn.zero_grad()
	# input = torch.unsqueeze(input,0)
	input = Variable(input) # Trying to remember the math but memory was being shared amoungst varible
	output = cnn(input)
	loss = criterion(output, Variable(torch.from_numpy(target).type(torch.DoubleTensor)))
	loss.backward(retain_graph=True) #Gradient
	optimizer.step() # Backprop		
	print("loss: "+str(loss.data[0]))#+"\t output: "+str(output.data[0,0])+"\tTarget: "+str(getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[start+32]))

print("\n\n\n")
print("Printing Predictions")
#Launching Predictions
for day in range(trainingData,lengthOfCloseData-34):
	input = matrix[:,day:day+32]
	input = torch.from_numpy(input).type(torch.DoubleTensor)
	input = torch.unsqueeze(input,0)
	input = Variable(input) # Trying to remember the math but memory was being shared amoungst varibles
	output = cnn(input)
	loss = criterion(output, Variable(torch.DoubleTensor([getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day+32]])))
	print("loss: "+str(loss.data[0])+"\t output: "+str(output.data[0,0])+"\tTarget: "+str(getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day+32]))	


