from __future__ import unicode_literals, print_function, division
from io import open
import glob
import unicodedata
import string
import torch
import torch.nn as nn
from torch.autograd import Variable
import openpyxl
import random
import torch.nn.functional as F
from matplotlib import pyplot as plt
import numpy as np

#Double check the percent change is doing what it is supposed to - Done
#Summary Metric for test accuracy, avg loss over test set and graph of real and predictions.
#Get avergage of loss / |correct label| - Done
#Parameterize the code how many days of training and test. Offset Day(Number of days) - Done
#What do open and close even mean for bitcoin (After hours trading stocks?) - Done (It is stupid 12 UTC)

#TODO Write Function that reverses the order of work because most recent data is first
# Lists behave like stack
workbook = openpyxl.load_workbook('/Users/calvinusiri/Desktop/CS/machinelearning/cryptoPrice1.xlsx')
type(workbook)
dictionaryOfCryptoObjects = dict()
sheetNames = ["BTC","ETH","XRP","LTC"]
crypto = "LTC"
class Crypto(object):
	open = []
	high = []
	low = []
	close = []
	volume = []
	mktCap = []
	date = []
	def __init__(self,open,high,low,close,volume,mktCap,date):
		super(Crypto, self).__init__()
		self.open = open
		self.high = high
		self.low = low
		self.close = close
		self.volume = volume
		self.mktCap = mktCap
		self.date = date

class RNN(nn.Module):
	"""docstring for RNN"""
	def __init__(self, input_size,hidden_size,output_size):
		super(RNN, self).__init__()
		self.hidden_size = hidden_size
		self.i2h = nn.Linear(input_size + hidden_size, hidden_size)
		self.i2o = nn.Linear(input_size + hidden_size, output_size)
		
	def forward(self,input,hidden):
		input_combined = torch.cat((input,hidden),1)
		hidden = F.sigmoid(self.i2h(input_combined))# essentially activation function
		output = self.i2o(input_combined)
		return output, hidden

	def initHidden(self):
		return Variable(torch.zeros(1, self.hidden_size))

# To get number between -1 and 1
def normalizeData(array,curr):
	return (curr-min(array))/(max(array)-min(array))

def getPercentageChange(arrays):
	container = []
	for a, b in zip(arrays[::1], arrays[1::1]):
		container.append(((b - a) / a)) # Removed 100* because, we want numbers close to 0 for sigmoid function
	return container

def reverseList(arr):
	arr.reverse()
	return arr

for i in sheetNames:
	sheet = workbook.get_sheet_by_name(i)
	open = []
	high = []
	low = []
	close = []
	volume = []
	mktCap = []
	date = []
	for j in range(1,sheet.max_column+1):
		value = sheet.cell(row=1,column=j).value
		if value == "Date":
			for k in range(2,sheet.max_row+1):
				date.append(sheet.cell(row=k,column=j).value)
		if value == "Open":
			for k in range(2,sheet.max_row+1):
				open.append(sheet.cell(row=k,column=j).value)
		if value == "High":
			for k in range(2,sheet.max_row+1):
				high.append(sheet.cell(row=k,column=j).value)
		if value == "Low":
			for k in range(2,sheet.max_row+1):
				low.append(sheet.cell(row=k,column=j).value)
		if value == "Close":
			for k in range(2,sheet.max_row+1):
				close.append(sheet.cell(row=k,column=j).value)
		if value == "Volume":
			for k in range(2,sheet.max_row+1):
				value = sheet.cell(row=k,column=j).value
				if value == "-":
					volume.append(0.5)
				else:
					volume.append(value)
		if value == "Market Cap":
			for k in range(2,sheet.max_row+1):
				mktCap.append(sheet.cell(row=k,column=j).value)
	dictionaryOfCryptoObjects[i] = Crypto(reverseList(open),reverseList(high),reverseList(low),reverseList(close),reverseList(volume),reverseList(mktCap),reverseList(date))

#Network
lengthOfCloseData = dictionaryOfCryptoObjects[crypto].close
trainingData = int(len(lengthOfCloseData)*0.80)
testingData = int(len(lengthOfCloseData) - trainingData)
numberOfDaysToPredict = 1
offset = numberOfDaysToPredict + 1
lossArray = []
PredictionArray = []
ActualArray = []
dateArray = []

print("Training:",trainingData)
print("Testing:",testingData)

#Training The Network
rnn = RNN(6, 32, 1)
hidden = rnn.initHidden()
criterion = nn.L1Loss()
learning_rate = 0.005 # Needs to be just right, if its too high then it might explode if its too low then it might not learn, no need because we have optim
optim = torch.optim.SGD(rnn.parameters(),lr = learning_rate)
hidden = Variable(torch.zeros(1,32)) #row col  hidden should not be within the learning step

for day in range(trainingData):
	rnn.zero_grad() # Want to reset the gradient ever learning step
	input = torch.zeros(1,6)
	input[0][0] = getPercentageChange(dictionaryOfCryptoObjects[crypto].open)[day]
	input[0][1] = getPercentageChange(dictionaryOfCryptoObjects[crypto].high)[day]
	input[0][2] = getPercentageChange(dictionaryOfCryptoObjects[crypto].low)[day]
	input[0][3] = getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day]
	# input[0][4] = getPercentageChange(dictionaryOfCryptoObjects[crypto].volume)[day]
	input[0][4] = 0.5
	input[0][5] = getPercentageChange(dictionaryOfCryptoObjects[crypto].mktCap)[day]
	input = Variable(input) # Trying to remember the math but memory was being shared amoungst varibles
	output,hidden = rnn(input,hidden)
	loss = criterion(output, Variable(torch.FloatTensor([getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day+numberOfDaysToPredict]])))
	loss.backward(retain_graph=True) #Gradient
	optim.step() # Backprop
	# print("Date Testing:",dictionaryOfCryptoObjects[crypto].date[day])

# Using the model to predict and see how right we are
for day in range(trainingData, (trainingData+testingData)-offset):
	input = torch.zeros(1,6)
	input[0][0] = getPercentageChange(dictionaryOfCryptoObjects[crypto].open)[day]
	input[0][1] = getPercentageChange(dictionaryOfCryptoObjects[crypto].high)[day]
	input[0][2] = getPercentageChange(dictionaryOfCryptoObjects[crypto].low)[day]
	input[0][3] = getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day]
	# input[0][4] = getPercentageChange(dictionaryOfCryptoObjects[crypto].volume)[day]
	input[0][4] = 0.5
	input[0][5] = getPercentageChange(dictionaryOfCryptoObjects[crypto].mktCap)[day]
	input = Variable(input)
	output = rnn(input,hidden)
	print("---------------------------------")
	print("Date Prediction:",dictionaryOfCryptoObjects[crypto].date[day])
	print("Modal Prediction:",output[0].data[0][0])
	print("Actual Result:",getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day+numberOfDaysToPredict])
	loss = criterion(output[0], Variable(torch.FloatTensor([getPercentageChange (dictionaryOfCryptoObjects[crypto].close)[day+numberOfDaysToPredict]])))
	print("Loss Average Absolute:",loss.data[0]/abs(0.00001 + getPercentageChange (dictionaryOfCryptoObjects[crypto].close)[day]))
	print("Loss Function Results:",loss.data[0])
	# ActualArray.append(getPercentageChange(dictionaryOfCryptoObjects[crypto].close)[day])
	# PredictionArray.append(output[0])
	# lossArray.append(loss)
	# dateArray.append(dictionaryOfCryptoObjects[crypto].date[day])